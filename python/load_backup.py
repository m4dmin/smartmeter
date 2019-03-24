#!/usr/bin/env python3

import configparser
import logging
from logging.handlers import RotatingFileHandler
import sys
from influxdb import InfluxDBClient
from openpyxl import load_workbook


############### init section ###############

##### config section
config = configparser.ConfigParser()
config.read('/smartmeter/conf/load_backup.conf')

influxDB_ip = config['influxDB']['ipAdresse']
influxDB_port = config['influxDB']['port']
influxDB_user = config['influxDB']['user']
influxDB_passwd = config['influxDB']['password']
influxDB_db = config['influxDB']['db']
influxDB_tag_instance = config['influxDB']['tag_instance']
influxDB_tag_source = config['influxDB']['tag_source']

backup_dsn = config['backup']['dsn']

##### Array with JSON data for influxDB write
points = []

##### measurement section

msmntDict = {"sz_leistung_aktuell": "W",
             "pv_leistung_aktuell": "W",
             "verbrauch_aktuell": "W",
             "sz_tariflos_1.8.0": "kWh",
             "sz_hochtarif_1.8.1": "kWh",
             "sz_niedertarif_1.8.2": "kWh",
             "sz_einspeisen_tariflos_2.8.0": "kWh",
             "sz_einspeisen_hochtarif_2.8.1": "kWh",
             "sz_einspeisen_niedertarif_2.8.2": "kWh",
             "pv_ertrag_tag": "kWh",
             "einspiesen_tag": "kWh",
             "bezug_tag": "kWh",
             "pv_eigenvergrauch_tag": "kWh",
             "verbrauch_tag": "kWh"
            }   

############### logging section ###############
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# create a file handler
handler = RotatingFileHandler('/smartmeter/log/smartmeter.log', maxBytes=10*1024*1024, backupCount=2)
handler.setLevel(logging.INFO)

# create a logging format
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)

# add the handlers to the logger
logger.addHandler(handler)

# logging examples
# logger.debug("Debug Log")
# logger.info("Info Log")
# logger.warning("Warning Log")
# logger.error("Error Log")
# logger.critical("Critical Log")


############### Runtime section ##################

# Excel Workbook auslesen und in JSON umwandeln
try:
    logger.info("Lade Daten in DB "+influxDB_db+", Instanz "+influxDB_tag_instance+", Source "+influxDB_tag_instance)
    logger.info("Backup-Datei: "+str(backup_dsn))
    workbook = load_workbook(str(backup_dsn))
    sheetnames = workbook.sheetnames
except Exception as e:
    logger.error(e)
    sys.exit(1)

try:
    for sheet in sheetnames:
        worksheet = workbook.get_sheet_by_name(sheet)
        for row in worksheet.values:
            logger.debug("SHEET: "+sheet)
            logger.debug("TIME:  "+row[0])
            logger.debug("VALUE: "+row[1])

            if row[0] == "Time":
                continue

            logger.info("Lade Datensatz: measurement "+sheet+", time "+row[0]+", value "+row[1])

            for msmnt, msmnt_unit in msmntDict.items():
                if sheet == msmnt:
                    unit = msmnt_unit
            point = {
                "measurement": sheet,
                "tags": {
                    "instance": influxDB_tag_instance,
                    "source": influxDB_tag_source
                },
                "time": row[0],
                "fields": {
                    unit: row[1]
                    }
                }
            points.append(point)
            logger.debug(point)

except Exception as e:
    logger.error(e)
    sys.exit(1)


##### influxDB section
try:
    clientInfluxDB = InfluxDBClient(influxDB_ip, influxDB_port, influxDB_user, influxDB_passwd, influxDB_db)
    clientInfluxDB.write_points(points)
except Exception as e:
    logger.error(e)
    sys.exit(1)